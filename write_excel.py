
import openpyxl


class Excel_Template:
    def __init__(self, cols, col_names, sheet_name):
        self.num_cols = cols
        self.data = []
        self.wb = openpyxl.Workbook()

        # for name in self.wb.get_sheet_names():
        #     print name
        #     self.wb.remove_sheet(name)
        self.wb.create_sheet(index=0, title=sheet_name)

        self.write_row(1, col_names, sheet_name)

        print self

    def make_sheet(self, name):
        self.wb.create_sheet(name)

    def write_row(self, num, data, sheet_name="data"):
        row_num = num
        col_num = 1

        try:
            sheet = self.wb.get_sheet_by_name(sheet_name)
        except:
            print "Making new sheet", sheet_name
            sheet = self.wb.create_sheet(sheet_name)

        for x in range(0, data.__len__()):
            sheet.cell(row=row_num, column=x+1).value = data[x]
            # print data[x]



    def save(self, path):
        self.wb.save(path)


# names = ["First Name", "Last Name", " Eye Color", "Balance", "Company"]
# t = Excel_Template(names.__len__(), names)


